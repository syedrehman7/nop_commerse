import openpyxl


def numRows(fileName, sheetName):
    Excel_file = openpyxl.load_workbook(fileName)
    Sheet = Excel_file[sheetName]
    return Sheet.max_row

def readData(fileName, sheetName, rowNum, colNum):
    Excel_file = openpyxl.load_workbook(fileName)
    Sheet = Excel_file[sheetName]
    return Sheet.cell(row=rowNum, column=colNum).value

def writeData(fileName, sheetName, rowNum, colNum, data):
    Excel_file = openpyxl.load_workbook(fileName)
    Sheet = Excel_file[sheetName]
    Sheet.cell(row=rowNum, column=colNum).value = data
    Excel_file.save(fileName)






